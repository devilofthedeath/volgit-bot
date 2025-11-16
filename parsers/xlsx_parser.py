import openpyxl
from openpyxl.styles import PatternFill
from core.models import Event, Schedule

class WorkingParser:
    def __init__(self):
        self.color_map = {
            "FF00FF00": "Неделя инноваций",      # зеленый
            "FF00FFFF": "Волга-IT",              # голубой  
            "FFFFFF00": "Студенческие СКБ",      # желтый
            "FFFF00FF": "Молодёжные метавселенные" # фиолетовый
        }
        self.sheet_structure = {
            "18.11": {
                "sheet_name": "РЕГЛАМЕНТ на 18.11",
                "time_col": 1,
                "locations_row": 1,
                "first_data_row": 2,
                "default_event_type": "Неделя инноваций"
            },
            "19.11": {
                "sheet_name": "РЕГЛАМЕНТ на 19.11", 
                "time_col": 1,
                "locations_row": 1,
                "first_data_row": 2,
                "default_event_type": "Волга-IT"
            },
            "20.11": {
                "sheet_name": "РЕГЛАМЕНТ на 20.11",
                "time_col": 1, 
                "locations_row": 1,
                "first_data_row": 2,
                "default_event_type": "Волга-IT"
            }
        }
    
    def parse_file(self, file_path: str):
        print("🚀 ЗАПУСК ПАРСЕРА (РАЗРЕШАЕМ ПАРАЛЛЕЛЬНЫЕ СОБЫТИЯ)")
        print("=" * 70)
        
        workbook = openpyxl.load_workbook(file_path)
        all_events = []
        
        for date, structure in self.sheet_structure.items():  # ← ТЕПЕРЬ ЕСТЬ!
            if structure["sheet_name"] in workbook.sheetnames:
                print(f"\n📋 ПАРСИМ {date}: {structure['sheet_name']}")
                sheet = workbook[structure["sheet_name"]]
                
                # Получаем места с диагностикой дубликатов
                locations = self._get_locations(sheet)
                
                events = self._parse_sheet_comprehensive(sheet, date, structure, locations)
                all_events.extend(events)
                print(f"   ✅ Найдено событий: {len(events)}")
        
        print(f"\n🎉 ВСЕГО СОБЫТИЙ: {len(all_events)}")
        
        # Возвращаем Schedule с last_updated
        from datetime import datetime
        return Schedule(
            events=all_events,
            last_updated=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
    
    def _parse_sheet(self, sheet, date: str) -> list[Event]:
        """Парсинг одного листа"""
        events = []
        
        # Сначала получим список мест из первой строки
        locations = self._get_locations(sheet)
        print(f"   Места проведения: {locations}")
        
        # Проходим по всем строкам (начиная со второй, т.к. первая - заголовки)
        for row_idx in range(2, sheet.max_row + 1):
            time_cell = sheet.cell(row_idx, 1)  # Первый столбец - время
            if not time_cell.value:
                continue
                
            time_str = str(time_cell.value).strip()
            print(f"⏰ Анализируем время: {time_str}")
            
            # Проходим по всем столбцам с местами
            for col_idx in range(2, len(locations) + 2):
                cell = sheet.cell(row_idx, col_idx)
                
                # Если ячейка не пустая - создаем событие
                if cell.value and str(cell.value).strip():
                    location = locations[col_idx - 2]  # -2 т.к. начинаем с колонки 2
                    title = str(cell.value).strip()
                    
                    # Определяем тип события по цвету
                    event_type = self._get_event_type(cell)
                    color = self._get_color_name(cell.fill)
                    
                    # Пока используем фиксированную продолжительность 30 мин
                    # Позже добавим логику для объединенных ячеек
                    start_time = time_str
                    end_time = self._calculate_end_time(time_str, 30)
                    
                    event = Event(
                        date=date,
                        start_time=start_time,
                        end_time=end_time,
                        location=location,
                        title=title,
                        event_type=event_type,
                        color=color
                    )
                    
                    events.append(event)
                    print(f"     ✅ Событие: {time_str} | {location} | {title}")
        
        return events
    
    def _get_locations(self, sheet) -> list[str]:
        """Получаем список мест из первой строки"""
        locations = []
        # Вторая строка (индекс 2) содержит названия мест
        for col_idx in range(2, sheet.max_column + 1):
            cell = sheet.cell(2, col_idx)
            if cell.value:
                locations.append(str(cell.value).strip())
        return locations
    
    def _get_event_type(self, cell) -> str:
        """Определяем тип события по цвету ячейки"""
        if cell.fill and cell.fill.fgColor:
            color_hex = cell.fill.fgColor.rgb
            return self.color_map.get(color_hex, "Другое")
        return "Другое"
    
    def _get_color_name(self, fill) -> str:
        """Получаем название цвета"""
        if fill and fill.fgColor:
            color_hex = fill.fgColor.rgb
            color_names = {
                "FF00FF00": "green",
                "FF00FFFF": "blue", 
                "FFFFFF00": "yellow",
                "FFFF00FF": "purple"
            }
            return color_names.get(color_hex, "unknown")
        return "none"
    
    def _calculate_end_time(self, start_time: str, duration_minutes: int) -> str:
        """Вычисляем время окончания (упрощенная версия)"""
        # Пока просто добавляем 30 минут
        # Позже добавим логику для нерегулярного времени
        return "30_min_later"  # Заглушка
    
    def _parse_sheet_advanced(self, sheet, date: str, time_col: int, locations_row: int, locations: list) -> list[Event]:
        """Улучшенный парсинг с обработкой объединенных ячеек"""
        events = []
        
        # Получаем все объединенные ячейки
        merged_ranges = list(sheet.merged_cells.ranges) if sheet.merged_cells else []
        
        # Проходим по всем строкам после заголовка мест
        for row_idx in range(locations_row + 1, sheet.max_row + 1):
            time_cell = sheet.cell(row_idx, time_col)
            if not time_cell.value or not ":" in str(time_cell.value):
                continue
                
            time_str = str(time_cell.value).split()[0]  # Берем только время, убираем дату если есть
            print(f"   ⏰ Время: {time_str}")
            
            # Проходим по всем колонкам с местами
            for col_idx in range(time_col + 1, time_col + 1 + len(locations)):
                if col_idx > sheet.max_column:
                    continue
                    
                cell = sheet.cell(row_idx, col_idx)
                location_index = col_idx - (time_col + 1)
                location = locations[location_index] if location_index < len(locations) else f"Место_{col_idx}"
                
                # Проверяем, является ли ячейка частью объединенного диапазона
                is_merged, merged_value, merged_height = self._check_merged_cell(merged_ranges, row_idx, col_idx, sheet)
                
                if is_merged:
                    # Это объединенная ячейка - берем значение из первой ячейки диапазона
                    cell_value = merged_value
                    duration_minutes = merged_height * 30  # Предполагаем шаг 30 минут
                else:
                    cell_value = cell.value
                    duration_minutes = 30
                
                # Если ячейка не пустая - создаем событие
                if cell_value and str(cell_value).strip():
                    title = str(cell_value).strip()
                    
                    # Определяем тип события по цвету
                    event_type = self._get_event_type(cell)
                    color = self._get_color_name(cell.fill)
                    
                    # Вычисляем время окончания
                    end_time = self._calculate_real_end_time(time_str, duration_minutes)
                    
                    event = Event(
                        date=date,
                        start_time=time_str,
                        end_time=end_time,
                        location=location,
                        title=title,
                        event_type=event_type,
                        color=color
                    )
                    
                    events.append(event)
                    print(f"     ✅ {time_str}-{end_time} | {location} | {title} | {event_type}")
        
        return events

    def _check_merged_cell(self, merged_ranges, row_idx, col_idx, sheet):
        """Проверяет, является ли ячейка частью объединенного диапазона"""
        for merged_range in merged_ranges:
            if (row_idx >= merged_range.min_row and row_idx <= merged_range.max_row and
                col_idx >= merged_range.min_col and col_idx <= merged_range.max_col):
                
                # Берем значение из первой ячейки диапазона
                first_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                height = merged_range.max_row - merged_range.min_row + 1
                
                return True, first_cell.value, height
        
        return False, None, 0
    def _parse_sheet_comprehensive(self, sheet, date: str, structure: dict, locations: list):
        """Парсим ВСЕ события, разрешаем параллельные мероприятия"""
        events = []
        time_col = structure["time_col"]
        first_data_row = structure["first_data_row"]
        
        # Проходим по всем строкам после заголовка мест
        for row_idx in range(first_data_row, sheet.max_row + 1):
            time_cell = sheet.cell(row_idx, time_col)
            if not time_cell.value or not ":" in str(time_cell.value):
                continue
                
            time_str = str(time_cell.value).split()[0]  # Берем только время
            
            for col_idx in range(2, 2 + len(locations)):
                if col_idx > sheet.max_column:
                    continue
                    
                cell = sheet.cell(row_idx, col_idx)
                location_idx = col_idx - 2
                location = locations[location_idx] if location_idx < len(locations) else f"Место_{col_idx}"
                
                if cell.value and str(cell.value).strip():
                    title = str(cell.value).strip()
                    
                    # Определяем тип события по цвету
                    event_type = self._get_event_type(cell)
                    color_hex = self._get_color_hex(cell.fill)
                    
                    # Если не определили по цвету, используем тип по умолчанию для листа
                    if event_type == "Другое":
                        event_type = structure["default_event_type"]
                    
                    # Пока используем фиксированную продолжительность 30 мин
                    start_time = time_str
                    end_time = self._calculate_end_time(time_str, 30)
                    
                    event = Event(
                        date=date,
                        start_time=start_time,
                        end_time=end_time,
                        location=location,
                        title=title,
                        event_type=event_type,
                        color=color_hex
                    )
                    
                    events.append(event)
                    print(f"     ✅ {time_str}-{end_time} | {location:20} | {title:25} | {event_type}")
        
        return events
    def _get_color_hex(self, fill):
        """Получаем hex цвета"""
        if fill and fill.fgColor:
            return fill.fgColor.rgb
        return None
    def _calculate_real_end_time(self, start_time: str, duration_minutes: int) -> str:
        """Реально вычисляем время окончания"""
        from datetime import datetime, timedelta
        
        try:
            # Парсим время (может быть в формате "9:00", "09:00", "9:00:00")
            time_formats = ["%H:%M", "%H:%M:%S", "%H.%M"]
            time_obj = None
            
            for fmt in time_formats:
                try:
                    time_obj = datetime.strptime(start_time, fmt)
                    break
                except ValueError:
                    continue
            
            if time_obj:
                end_time = time_obj + timedelta(minutes=duration_minutes)
                return end_time.strftime("%H:%M")
            else:
                return f"+{duration_minutes}min"
        except:
            return f"+{duration_minutes}min"